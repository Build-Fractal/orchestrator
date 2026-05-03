#!/usr/bin/env python3
"""xlsx-to-csv.py -- Tier 1 XLSX -> per-sheet CSV shim.

Usage: xlsx-to-csv.py <input.xlsx> --out-dir <target-dir>

Emits one CSV per sheet under <target-dir>, named <sanitized-sheet>.csv,
with the first row treated as the header (per US-6 AS-3 / SC-9). Uses
openpyxl in read_only mode for memory efficiency on large regulatory
sheets. Exits 2 if openpyxl is not installed (with install hint on
stderr); exits 1 on bad arguments or missing input.

Stdout: one `CSV: <abs-path>` line per emitted file, plus a final
`SUMMARY: xlsx-to-csv sheets=<N>` line.
"""
import argparse
import csv
import os
import re
import sys

try:
    import openpyxl
except ImportError:
    sys.stderr.write(
        "xlsx-to-csv.py: openpyxl not installed; "
        "install via 'pipx install openpyxl' or "
        "'python3 -m pip install --user openpyxl'\n"
    )
    sys.exit(2)


def sanitize_sheet_name(name):
    """Sanitize a sheet name for use as a filename.

    Replace runs of `/` or whitespace with `-`, then strip leading/trailing
    dashes. Deterministic so the verifier's `diff -q` against expected
    files is reliable.
    """
    sanitized = re.sub(r"[/\s]+", "-", name).strip("-")
    return sanitized or "sheet"


def main(argv):
    parser = argparse.ArgumentParser(
        description="XLSX -> per-sheet CSV (header = first row).",
    )
    parser.add_argument("input", help="path to input .xlsx file")
    parser.add_argument(
        "--out-dir",
        required=True,
        help="directory to write per-sheet CSVs into (created if missing)",
    )
    args = parser.parse_args(argv)

    if not os.path.isfile(args.input):
        sys.stderr.write(
            "xlsx-to-csv.py: input not found: %s\n" % args.input
        )
        return 1

    os.makedirs(args.out_dir, exist_ok=True)

    wb = openpyxl.load_workbook(args.input, data_only=True, read_only=True)
    count = 0
    for name in wb.sheetnames:
        ws = wb[name]
        sanitized = sanitize_sheet_name(name)
        out_path = os.path.abspath(os.path.join(args.out_dir, sanitized + ".csv"))
        with open(out_path, "w", newline="") as fh:
            writer = csv.writer(fh)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(["" if cell is None else cell for cell in row])
        sys.stdout.write("CSV: %s\n" % out_path)
        count += 1
    sys.stdout.write("SUMMARY: xlsx-to-csv sheets=%d\n" % count)
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
